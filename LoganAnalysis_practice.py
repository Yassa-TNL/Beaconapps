'''
*****************************************************************#
Title                            Logan Anaylsis
Purpose                          Complies dataset from
                                 multiple log files
                                 and last backup of each month
Creation Date                    08/27/2020
Created by                       John
Modified by                      Derek Vincent Taylor, Jason R. Bock
Last Modified                    06/23/2025
*****************************************************************
In progress                     1. Add recognition score hitrate - false alarm
                                Target-%Cor - Foil-%Inc
                                2. Caclulate d' values handle 0 or 1 hit rate
                                or false alarm
                                3. d' for lures value d' equation
                                {Target-%Cor ,LureH-%Inc, LureL-%inc} see clack
                                see normsinv function equation
                                4. Run script on cron job on server

'''

from __future__ import division
import os
from openpyxl import Workbook
import numpy as np
from numpy import trapz
import sys


def main():
    error_log = "Data Parsed, Error Occured During, notes" + "\n"
    # to see what each block of code is doing examine error_level statement
    error_level = ""
    try:

        # Create spreadsheet, and 1 sheet for each task type
        sheetName = "mdt_results.xlsx"
        wb = Workbook()
        wsObj = wb.active
        wsObj.title = "MDTO"
        wsSpl = wb.create_sheet(title="MDTS")
        wsTmp = wb.create_sheet(title="MDTT")
        print("")

        # Assign the column headers for each sheet
        error_log = "Assigning Workbooks, Formatting worksheets in Workbook" + "\n"
        for ws in wb:
            try:
                if (ws.title == "MDTO"):
                    error_level = "Assigning Workbooks, Formatting MDTO worksheet in Workbook" + "\n"
                    cdHard = "Target"
                    cdHigh = "LureH"
                    cdLow = "LureL"
                    cdEasy = "Foil"
                elif (ws.title == "MDTS"):
                    error_level = "Assigning Workbooks, Formatting MDTS worksheet in Workbook" + "\n"
                    cdHard = "Same"
                    cdHigh = "Small Mv"
                    cdLow = "Large Mv"
                    cdEasy = "Corner Mv"
                elif (ws.title == "MDTT"):
                    error_level = "Assigning Workbooks, Formatting MDTT worksheet in Workbook" + "\n"
                    cdHard = "Adj"
                    cdHigh = "Eight"
                    cdLow = "Sixteen"
                    cdEasy = "PR"
                error_level = "Assigning Workbooks, Column Titles Workbook" + "\n"
                ws['A1'] = "Subject"
                ws['B1'] = "Trials/Cond"
                ws['C1'] = "%s-Responses" % (cdHard)
                ws['D1'] = "%s-%%Cor" % (cdHard)
                ws['E1'] = "%s-%%Inc" % (cdHard)
                ws['F1'] = "%s-Responses" % (cdHigh)
                ws['G1'] = "%s-%%Cor" % (cdHigh)
                ws['H1'] = "%s-%%Inc" % (cdHigh)
                ws['I1'] = "%s-Responses" % (cdLow)
                ws['J1'] = "%s-%%Cor" % (cdLow)
                ws['K1'] = "%s-%%Inc" % (cdLow)
                ws['L1'] = "%s-Responses" % (cdEasy)
                ws['M1'] = "%s-%%Cor" % (cdEasy)
                ws['N1'] = "%s-%%Inc" % (cdEasy)
            except Exception as e:
                error_log = error_log + error_level + ',' + str(sys.exc_info()[1]) + '\n'
        error_level = "Create Worksheet, MOTO-LDI workseet" + "\n"
        wsOLDI = wb.create_sheet(title="MDTO-LDI")
        wsOLDI['A1'] = "Subject"
        wsOLDI['B1'] = "d'"
        wsOLDI['C1'] = "LDI: High"
        wsOLDI['D1'] = "LDI: Low"
        wsOLDI['E1'] = "LDI: Combined"
        wsOLDI['F1'] = "LDI_AUC"
        wsOLDI['G1'] = "Target-Foil_AUC"
        wsOLDI['H1'] = "LDI_slope"
        wsOLDI['I1'] = "Target-Foil_slope"

        error_level = "Create Worksheet, MDTS-LDI workseet" + "\n"
        wsSLDI = wb.create_sheet(title="MDTS-LDI")
        wsSLDI['A1'] = "Subject"
        wsSLDI['B1'] = "d'"
        wsSLDI['C1'] = "LDI: High"
        wsSLDI['D1'] = "LDI: Low"
        wsSLDI['E1'] = "LDI: Combined"
        wsSLDI['F1'] = "LDI_AUC"
        wsSLDI['G1'] = "Target-Foil_AUC"
        wsSLDI['H1'] = "LDI_slope"
        wsSLDI['I1'] = "Target-Foil_slope"

        error_level = "Create Worksheet, MDTT-LDI workseet" + "\n"
        # Analogous to above
        # Put all log files in directory into list
        dataDir = os.getcwd()
        fileList = os.listdir(dataDir)
        logs = [f for f in fileList if f.endswith("log.txt") and "old" not in f]

        objIdx = 0
        sptIdx = 0
        tmpIdx = 0
        error_level = "Processing Data, Iterating through logs," + "\n"
        for i in range(0, len(logs)):
            try:
                print(logs[i])
                error_level = "Processing Data, removing end of line characters," + "\n"
                logFile = [line.rstrip('\n') for line in open(logs[i])]
                error_level = "Processing Data, Testing precense of scores," + "\n"
                if not any(line.startswith("Scores:") for line in logFile):
                    print("No scores found, skipped logfile: %s" % (logs[i]))
                    continue
                error_level = "Processing Data, Creting Row," + "\n"
                row = ""
                error_level = "Processing Data, Getting Subject number," + "\n"
                subjectNum = int(logs[i].split("_")[0])
                error_level = "Processing Data, getting task type," + "\n"
                taskType = logs[i].split("_")[1]
                if taskType == "MDTO":
                    objIdx += 1
                    row = objIdx
                elif taskType == "MDTS":
                    sptIdx += 1
                    row = sptIdx
                elif taskType == "MDTT":
                    tmpIdx += 1
                    row = tmpIdx
                ws = wb[taskType]
                error_level = "Processing Data, set score index to zeero," + "\n"
                scoreIdx = 0
                error_level = "Processing Data, Iterating through length of logfile," + "\n"
                for j in range(0, len(logFile)):
                    try:
                        error_level = "Processing Data,getting trials," + "\n"
                        if logFile[j].startswith("Trials/Condition"):
                            trials = int(logFile[j].split(":")[1].strip())
                        error_level = "Processing Data,getting trials," + "\n"
                        if logFile[j].startswith("Blocks ran"):
                            trials = int(logFile[j][-3:]) * 4
                        error_level = "Processing Data,getting scores," + "\n"
                        if logFile[j].startswith("Scores"):
                            ws['C%d' % (row + 1)] = int(logFile[j + 4].split(":")[1].strip())
                            ws['F%d' % (row + 1)] = int(logFile[j + 7].split(":")[1].strip())
                            ws['I%d' % (row + 1)] = int(logFile[j + 10].split(":")[1].strip())
                            ws['L%d' % (row + 1)] = int(logFile[j + 13].split(":")[1].strip())
                            ws['D%d' % (row + 1)] = float(logFile[j + 15][-5:])
                            ws['E%d' % (row + 1)] = float(logFile[j + 16][-5:])
                            ws['G%d' % (row + 1)] = float(logFile[j + 17][-5:])
                            ws['H%d' % (row + 1)] = float(logFile[j + 18][-5:])
                            ws['J%d' % (row + 1)] = float(logFile[j + 19][-5:])
                            ws['K%d' % (row + 1)] = float(logFile[j + 20][-5:])
                            ws['M%d' % (row + 1)] = float(logFile[j + 21][-5:])
                            ws['N%d' % (row + 1)] = float(logFile[j + 22][-5:])
                            ws['A%d' % (row + 1)] = subjectNum
                            ws['B%d' % (row + 1)] = trials
                    except Exception as e:
                        error_log = error_log + error_level + ',' + str(sys.exc_info()[1]) + ' ' + logs[i] + '\n'

                if taskType == "MDTO":

                    pTgtHit = ws['D%d' % (row + 1)].value
                    pFoilFA = ws['N%d' % (row + 1)].value
                    if pTgtHit == 1: pTgtHit = 0.9999999999
                    if not pTgtHit: pTgtHit = 0.0000000001
                    if pFoilFA == 1: pFoilFA = 0.9999999999
                    if not pFoilFA: pFoilFA = 0.0000000001
                    dPrime = "=NORMSINV(%.2f)-NORMSINV(%.2f)" % (pTgtHit, pFoilFA)

                    hiLDI = ws['G%d' % (row + 1)].value - ws['E%d' % (row + 1)].value
                    loLDI = ws['J%d' % (row + 1)].value - ws['E%d' % (row + 1)].value
                    cbLDI = ((hiLDI * ws['F%d' % (row + 1)].value) + (loLDI * ws['I%d' % (row + 1)].value)) / (ws['F%d' % (row + 1)].value + ws['I%d' % (row + 1)].value)
                    yLDI = [loLDI, hiLDI]
                    areaLDI = trapz(yLDI, dx=1)
                    yTarFoil = [pTgtHit, pFoilFA]
                    areaTarFoil = trapz(yTarFoil, dx=1)
                    slopeLDI = loLDI - hiLDI
                    slopeTarFoil = pTgtHit - pFoilFA

                    wsOLDI['A%d' % (row + 1)] = subjectNum
                    wsOLDI['B%d' % (row + 1)] = dPrime
                    wsOLDI['C%d' % (row + 1)] = hiLDI
                    wsOLDI['D%d' % (row + 1)] = loLDI
                    wsOLDI['E%d' % (row + 1)] = cbLDI
                    wsOLDI['F%d' % (row + 1)] = areaLDI
                    wsOLDI['G%d' % (row + 1)] = areaTarFoil
                    wsOLDI['H%d' % (row + 1)] = slopeLDI
                    wsOLDI['I%d' % (row + 1)] = slopeTarFoil

                if taskType == "MDTS":

                    pTgtHit = ws['D%d' % (row + 1)].value
                    pFoilFA = ws['N%d' % (row + 1)].value
                    if pTgtHit == 1: pTgtHit = 0.9999999999
                    if not pTgtHit: pTgtHit = 0.0000000001
                    if pFoilFA == 1: pFoilFA = 0.9999999999
                    if not pFoilFA: pFoilFA = 0.0000000001
                    dPrime = "=NORMSINV(%.2f)-NORMSINV(%.2f)" % (pTgtHit, pFoilFA)

                    hiLDI = ws['G%d' % (row + 1)].value - ws['E%d' % (row + 1)].value
                    loLDI = ws['J%d' % (row + 1)].value - ws['E%d' % (row + 1)].value
                    cbLDI = ((hiLDI * ws['F%d' % (row + 1)].value) + (loLDI * ws['I%d' % (row + 1)].value)) / (ws['F%d' % (row + 1)].value + ws['I%d' % (row + 1)].value)
                    yLDI = [loLDI, hiLDI]
                    areaLDI = trapz(yLDI, dx=1)
                    yTarFoil = [pTgtHit, pFoilFA]
                    areaTarFoil = trapz(yTarFoil, dx=1)
                    slopeLDI = loLDI - hiLDI
                    slopeTarFoil = pTgtHit - pFoilFA
                    wsSLDI['A%d' % (row + 1)] = subjectNum
                    wsSLDI['B%d' % (row + 1)] = dPrime
                    wsSLDI['C%d' % (row + 1)] = hiLDI
                    wsSLDI['D%d' % (row + 1)] = loLDI
                    wsSLDI['E%d' % (row + 1)] = cbLDI
                    wsSLDI['F%d' % (row + 1)] = areaLDI
                    wsSLDI['G%d' % (row + 1)] = areaTarFoil
                    wsSLDI['H%d' % (row + 1)] = slopeLDI
                    wsSLDI['I%d' % (row + 1)] = slopeTarFoil
            except Exception as e:
                error_log = error_log + error_level + ',' + str(sys.exc_info()[1]) + ' ' + logs[i] + '\n'
        error_level = "Output file, sort each sheet by subject ID," + "\n"
        wb.save(filename=sheetName)

    except Exception as e:
        error_log = error_log + error_level + ',' + str(sys.exc_info()[1]) + '\n'
    print(error_log)


if __name__ == '__main__': main()




